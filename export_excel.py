"""
export_excel.py -- Xuất dữ liệu processed ra file Excel để quan sát trực quan.

Tạo file Excel với 4 sheet:
  1. Tổng quan   -- Thống kê tổng hợp (số bài, phân loại, top danh mục...)
  2. Tác giả     -- Danh sách bài type=author
  3. Tác phẩm    -- Danh sách bài type=work
  4. Khái niệm   -- Danh sách bài type=concept

Chạy:
    cd C:\\Users\\Admin\\Desktop\\ĐATN
    pip install openpyxl pandas
    python export_excel.py
"""

import json
from pathlib import Path
from datetime import datetime
from collections import Counter

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ============================================================
#  CAU HINH
# ============================================================
BASE_DIR   = Path(__file__).parent
PROC_DIR   = BASE_DIR / "data" / "processed"
OUTPUT     = BASE_DIR / "data" / "processed_data.xlsx"

# Mau sac theo type
COLOR_HEADER   = "1F497D"   # headerblue
COLOR_AUTHOR   = "B8D4EA"   # layerA (xanh nhạt)
COLOR_WORK     = "C6EFCE"   # rowgreen (xanh lá nhạt)
COLOR_CONCEPT  = "FFEB9C"   # rowyellow (vàng nhạt)
COLOR_SUMMARY  = "DCE6F1"   # rowblue

# Cot can xuat (khong xuat content day du - qua dai)
COLUMNS = [
    "id", "title", "type", "category",
    "word_count", "char_count",
    "summary_preview",          # 200 ký tự đầu của summary
    "categories_str",           # categories nối chuỗi
    "crawled_at", "last_modified", "processed_at",
    "url",
]

# ============================================================
#  ĐỌC DỮ LIỆU
# ============================================================

def load_processed_data():
    """Đọc tất cả file JSON trong data/processed/ → list of dict."""
    records = []
    for type_dir in ["author", "work", "concept"]:
        folder = PROC_DIR / type_dir
        if not folder.exists():
            continue
        for json_file in sorted(folder.glob("page_*.json")):
            try:
                with open(json_file, encoding="utf-8") as f:
                    rec = json.load(f)
                # Thêm trường tiện quan sát
                rec["summary_preview"] = (rec.get("summary", "") or "")[:200]
                cats = rec.get("categories", [])
                rec["categories_str"] = " | ".join(cats) if cats else ""
                records.append(rec)
            except Exception as e:
                print(f"  Lỗi đọc {json_file.name}: {e}")
    return records


def make_dataframe(records, doc_type=None):
    """Tạo DataFrame, lọc theo type nếu cần."""
    df = pd.DataFrame(records)
    if doc_type:
        df = df[df["type"] == doc_type].copy()
    # Chỉ giữ cột cần thiết, đúng thứ tự
    cols = [c for c in COLUMNS if c in df.columns]
    df = df[cols].reset_index(drop=True)
    df.index += 1   # Bắt đầu từ 1
    return df


# ============================================================
#  STYLE HELPERS
# ============================================================

def style_header_row(ws, row, bg_color, text_color="FFFFFF"):
    for cell in ws[row]:
        cell.font = Font(bold=True, color=text_color, size=10)
        cell.fill = PatternFill("solid", fgColor=bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right=Side(style="thin", color="CCCCCC"),
        )


def style_data_rows(ws, start_row, end_row, alt_color):
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row)):
        for cell in row:
            bg = alt_color if i % 2 == 1 else "FFFFFF"
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                bottom=Side(style="thin", color="E0E0E0"),
                right=Side(style="thin", color="E0E0E0"),
            )


def set_column_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def freeze_and_filter(ws, row=2):
    ws.freeze_panes = f"A{row}"
    ws.auto_filter.ref = ws.dimensions


# ============================================================
#  SHEET 1: TỔNG QUAN
# ============================================================

def write_summary_sheet(ws, records):
    ws.title = "Tổng quan"
    ws.sheet_view.showGridLines = False

    # --- Tiêu đề lớn ---
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "THỐNG KÊ DỮ LIỆU PROCESSED — Wikipedia Crawler ĐATN 2025-2026"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Thống kê chính ---
    total = len(records)
    by_type = Counter(r["type"] for r in records)
    by_cat  = Counter(r["category"] for r in records)
    words_all = [r.get("word_count", 0) for r in records]
    avg_words = int(sum(words_all) / total) if total else 0

    stats = [
        ("", "THỐNG KÊ TỔNG HỢP", ""),
        ("Tổng số bài",             total,                         "bài"),
        ("  Tác giả (author)",      by_type.get("author", 0),      "bài"),
        ("  Tác phẩm (work)",       by_type.get("work", 0),        "bài"),
        ("  Khái niệm (concept)",   by_type.get("concept", 0),     "bài"),
        ("Số từ trung bình / bài",  avg_words,                     "từ"),
        ("Bài nhiều từ nhất",       max(words_all) if words_all else 0, "từ"),
        ("Bài ít từ nhất",          min(words_all) if words_all else 0, "từ"),
        ("Thời gian export",        datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""),
    ]

    for i, (label, value, unit) in enumerate(stats, start=3):
        ws.cell(i, 1, label)
        ws.cell(i, 2, value)
        ws.cell(i, 3, unit)
        if label == "":
            ws.cell(i, 2).font = Font(bold=True, size=11, color=COLOR_HEADER)
            ws.merge_cells(f"B{i}:D{i}")
        else:
            ws.cell(i, 1).font = Font(size=10)
            ws.cell(i, 2).font = Font(bold=True, size=10)
            ws.cell(i, 3).font = Font(size=10, color="666666")
        bg = COLOR_SUMMARY if i % 2 == 0 else "FFFFFF"
        for c in [1, 2, 3]:
            ws.cell(i, c).fill = PatternFill("solid", fgColor=bg)

    # --- Top 15 danh mục ---
    row_offset = len(stats) + 5
    ws.cell(row_offset, 1, "TOP DANH MỤC (số bài crawl gốc)") \
      .font = Font(bold=True, size=11, color=COLOR_HEADER)

    for ci, (cat, cnt) in enumerate(by_cat.most_common(15), start=row_offset + 1):
        ws.cell(ci, 1, cat)
        ws.cell(ci, 2, cnt)
        bg = COLOR_SUMMARY if ci % 2 == 0 else "FFFFFF"
        ws.cell(ci, 1).fill = PatternFill("solid", fgColor=bg)
        ws.cell(ci, 2).fill = PatternFill("solid", fgColor=bg)
        ws.cell(ci, 2).alignment = Alignment(horizontal="right")

    # --- Biểu đồ phân loại ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "Số bài theo phân loại"
    chart.style = 10
    chart.y_axis.title = "Số bài"
    chart.x_axis.title = "Loại"

    # Dữ liệu biểu đồ (viết vào cột ẩn E, F)
    chart_data = [("Loại", "Số bài"),
                  ("author",  by_type.get("author", 0)),
                  ("work",    by_type.get("work", 0)),
                  ("concept", by_type.get("concept", 0))]
    for ci, (a, b) in enumerate(chart_data, start=3):
        ws.cell(ci, 5, a)
        ws.cell(ci, 6, b)

    data_ref   = Reference(ws, min_col=6, min_row=3, max_row=6)
    cats_ref   = Reference(ws, min_col=5, min_row=4, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws.add_chart(chart, "E10")

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10


# ============================================================
#  SHEET 2/3/4: DANH SÁCH BÀI
# ============================================================

COLUMN_LABELS = {
    "id":               "ID",
    "title":            "Tiêu đề",
    "type":             "Loại",
    "category":         "Danh mục crawl",
    "word_count":       "Số từ",
    "char_count":       "Số ký tự",
    "summary_preview":  "Tóm tắt (200 ký tự)",
    "categories_str":   "Danh mục Wikipedia",
    "crawled_at":       "Thời gian crawl",
    "last_modified":    "Sửa đổi cuối",
    "processed_at":     "Thời gian xử lý",
    "url":              "URL",
}

COLUMN_WIDTHS = {
    "A": 9,   # id
    "B": 32,  # title
    "C": 11,  # type
    "D": 25,  # category
    "E": 9,   # word_count
    "F": 11,  # char_count
    "G": 50,  # summary_preview
    "H": 38,  # categories_str
    "I": 20,  # crawled_at
    "J": 20,  # last_modified
    "K": 20,  # processed_at
    "L": 45,  # url
}


def write_data_sheet(ws, df, sheet_title, bg_color, alt_color):
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False

    # --- Header ---
    headers = [COLUMN_LABELS.get(c, c) for c in df.columns]
    for ci, h in enumerate(headers, start=1):
        ws.cell(1, ci, h)
    style_header_row(ws, 1, bg_color)
    ws.row_dimensions[1].height = 22

    # --- Dữ liệu ---
    for ri, row in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(ri, ci, val)
            # URL → hyperlink
            col_name = df.columns[ci - 1]
            if col_name == "url" and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    style_data_rows(ws, 2, ws.max_row, alt_color)
    set_column_widths(ws, COLUMN_WIDTHS)
    freeze_and_filter(ws)

    # Row height cho data
    for ri in range(2, ws.max_row + 1):
        ws.row_dimensions[ri].height = 55


# ============================================================
#  MAIN
# ============================================================

def main():
    print("=" * 55)
    print("  XUẤT DỮ LIỆU PROCESSED → EXCEL")
    print("=" * 55)

    if not PROC_DIR.exists():
        print("❌ Thư mục data/processed/ chưa tồn tại.")
        print("   Chạy python preprocess.py trước.")
        return

    print("📂 Đọc dữ liệu từ data/processed/...")
    records = load_processed_data()
    if not records:
        print("❌ Không tìm thấy dữ liệu nào trong data/processed/")
        return
    print(f"   ✅ Đọc được {len(records)} bài")

    # Tạo DataFrames
    df_all     = make_dataframe(records)
    df_author  = make_dataframe(records, "author")
    df_work    = make_dataframe(records, "work")
    df_concept = make_dataframe(records, "concept")

    print(f"   author:  {len(df_author)} bài")
    print(f"   work:    {len(df_work)} bài")
    print(f"   concept: {len(df_concept)} bài")

    # Ghi Excel bằng openpyxl trực tiếp
    print(f"\n📊 Tạo file Excel: {OUTPUT.name}...")

    # Dùng pandas writer trước để tạo structure, rồi style sau
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        # Ghi placeholder để tạo workbook
        pd.DataFrame().to_excel(writer, sheet_name="Tổng quan", index=False)
        df_author.to_excel(writer, sheet_name="Tác giả", index=True, index_label="STT")
        df_work.to_excel(writer, sheet_name="Tác phẩm", index=True, index_label="STT")
        df_concept.to_excel(writer, sheet_name="Khái niệm", index=True, index_label="STT")

    # Mở lại để style
    wb = load_workbook(OUTPUT)

    # Sheet 1: Tổng quan
    write_summary_sheet(wb["Tổng quan"], records)

    # Sheet 2: Tác giả
    ws_a = wb["Tác giả"]
    style_header_row(ws_a, 1, COLOR_HEADER)
    style_data_rows(ws_a, 2, ws_a.max_row, COLOR_AUTHOR)
    set_column_widths(ws_a, {get_column_letter(i+1): list(COLUMN_WIDTHS.values())[i]
                             for i in range(len(COLUMN_WIDTHS))})
    freeze_and_filter(ws_a)

    # Sheet 3: Tác phẩm
    ws_w = wb["Tác phẩm"]
    style_header_row(ws_w, 1, COLOR_HEADER)
    style_data_rows(ws_w, 2, ws_w.max_row, COLOR_WORK)
    set_column_widths(ws_w, {get_column_letter(i+1): list(COLUMN_WIDTHS.values())[i]
                             for i in range(len(COLUMN_WIDTHS))})
    freeze_and_filter(ws_w)

    # Sheet 4: Khái niệm
    ws_c = wb["Khái niệm"]
    style_header_row(ws_c, 1, COLOR_HEADER)
    style_data_rows(ws_c, 2, ws_c.max_row, COLOR_CONCEPT)
    set_column_widths(ws_c, {get_column_letter(i+1): list(COLUMN_WIDTHS.values())[i]
                             for i in range(len(COLUMN_WIDTHS))})
    freeze_and_filter(ws_c)

    wb.save(OUTPUT)

    print(f"\n✅ HOÀN THÀNH! File đã lưu tại:")
    print(f"   {OUTPUT}")
    print(f"\n   Nội dung:")
    print(f"   📋 Sheet 'Tổng quan'  — Thống kê + biểu đồ")
    print(f"   👤 Sheet 'Tác giả'    — {len(df_author)} bài")
    print(f"   📖 Sheet 'Tác phẩm'   — {len(df_work)} bài")
    print(f"   💡 Sheet 'Khái niệm'  — {len(df_concept)} bài")
    print("=" * 55)


if __name__ == "__main__":
    main()
